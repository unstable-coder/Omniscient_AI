from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

import xlrd
from openpyxl import load_workbook

from app.models.schemas import ContentUnit
from app.parsers.base_parser import BaseParser


class TabularParser(BaseParser):
    supported_extensions = ("csv", "tsv", "xlsx", "xls")

    supported_mime_types = (
        "text/csv",
        "text/tab-separated-values",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )

    def parse(
        self,
        path: str,
        metadata: dict[str, object],
    ) -> Iterable[ContentUnit]:

        suffix = Path(path).suffix.lower().lstrip(".")

        if suffix == "csv":
            yield from self._parse_delimited(path, ",", metadata)

        elif suffix == "tsv":
            yield from self._parse_delimited(path, "\t", metadata)

        elif suffix in ("xlsx", "xls"):
            yield from self._parse_excel(path, metadata)

    def _parse_delimited(
        self,
        path: str,
        delimiter: str,
        metadata: dict[str, object],
    ) -> Iterable[ContentUnit]:

        rows = []

        with open(
            path,
            newline="",
            encoding="utf-8",
            errors="ignore",
        ) as file:

            reader = csv.reader(file, delimiter=delimiter)

            for row in reader:

                values = [
                    str(cell).strip()
                    for cell in row
                    if str(cell).strip()
                ]

                if values:
                    rows.append(" | ".join(values))

        if rows:
            yield ContentUnit(
                text="\n".join(rows),
                metadata=metadata,
            )

    def _parse_excel(self, path: str, metadata: dict[str, object],) -> Iterable[ContentUnit]:
        suffix = Path(path).suffix.lower().lstrip(".")
        if suffix == "xlsx":
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=False,
            )
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    values = []
                    for cell in row:
                        if cell is None:
                            continue
                        value = str(cell).strip()
                        if value:
                            values.append(value)
                    if values:
                        rows.append(" | ".join(values))
                if rows:
                    yield ContentUnit(
                        text="\n".join(rows),
                        metadata={
                            **metadata,
                            "sheet_name": sheet.title,
                            "sheet_rows": len(rows),
                        },
                    )

            workbook.close()
        else:
            workbook = xlrd.open_workbook(path)
            for sheet in workbook.sheets():
                rows = []
                for r in range(sheet.nrows):
                    values = []
                    for cell in sheet.row_values(r):
                        value = str(cell).strip()
                        if value:
                            values.append(value)
                    if values:
                        rows.append(" | ".join(values))
                if rows:
                    yield ContentUnit(
                        text="\n".join(rows),
                        metadata={
                            **metadata,
                            "sheet_name": sheet.name,
                            "sheet_rows": len(rows),
                        },
                    )